# -*- coding: utf-8 -*-
"""
Created on Tue Sep 12 23:34:16 2023

@author: igors
"""

import pandas as pd
import copy
import math
import os
import re
import numpy as np
from openpyxl import load_workbook
import shutil
import random
from datetime import datetime


statistics_folder = r"C:\Users\igors\OneDrive\Dokumenty\\" 
excel_data = load_workbook(statistics_folder + "team_info.xlsx")
excel_data_sheet = excel_data["Data"]
year = excel_data_sheet['B1'].value
seriesname = excel_data_sheet['B2'].value
yourname = excel_data_sheet['B3'].value
results_path = excel_data_sheet['B4'].value
team_files_path = excel_data_sheet['B5'].value
rcd_file = excel_data_sheet['B6'].value
season = excel_data_sheet['B7'].value

def copying_orig_files():
    files_list = os.listdir(team_files_path)
    r = re.compile(".*.veh|.*.VEH")
    original_files = list(filter(r.match, files_list))
    if not os.path.exists(team_files_path + "Original"):
        os.mkdir(team_files_path + "Original")
    for file in original_files:
        shutil.copy(team_files_path + file, team_files_path + "Original" + "\\" + file)
    shutil.copy(rcd_file, team_files_path + "Original" + "\\" + "Original.rcd")

            

def find_name(text, start, end):
    """Finding name in xml file"""
    name_1 = text.index(start)
    name_2 = text.index(end)
    len_name_1 = len(start)
    return text[name_1 + len_name_1 : name_2]


def create_table(list_of_dfs):
    """Creating one dataframe from list of dataframes"""
    name_of_endtable = pd.DataFrame()
    name_of_endtable = list_of_dfs[0]
    for j in range(1,len(list_of_dfs)):
        name_of_endtable = name_of_endtable.merge(list_of_dfs[j], on=[0,1,2])
        

    name_of_endtable["Points Total"] = name_of_endtable.iloc[:, 1:].sum(axis = 1)
    name_of_endtable.sort_values("Points Total", ascending = False, inplace=True)
    name_of_endtable.reset_index(inplace=True)
    name_of_endtable.set_index(name_of_endtable.index+1, inplace=True)
    name_of_endtable.drop(columns="index", inplace=True)
    name_of_endtable.columns.values[0] = "Rider"
    name_of_endtable.columns.values[1] = "Team"
    name_of_endtable.columns.values[2] = "Car"
    for i in range(len(races)):
        name_of_endtable.columns.values[3+i] = race_tracks[0+i]
    return name_of_endtable
        
        
def create_team_table(table_with_riders):
    """Creating table for teams"""
    name_of_endtable = table_with_riders.iloc[:,1:]
    name_of_endtable = name_of_endtable.groupby(["Team", "Car"]).sum()
    name_of_endtable.sort_values("Points Total", ascending = False, inplace=True)
    name_of_endtable.reset_index(inplace=True)
    name_of_endtable.set_index(name_of_endtable.index+1, inplace=True)
    return name_of_endtable
    
def create_table_positions(table_with_riders):
    """Creating dataframe table with positions of riders"""
    name_of_endtable = table_with_riders.iloc[:,0:3]
    name_of_endtable.columns=[0,1,2]
    for j in range(0,len(table_df2)):
        name_of_endtable = name_of_endtable.merge(table_df2[j], on=[0,1,2])
    name_of_endtable.reset_index(inplace=True)
    name_of_endtable.set_index(table_with_riders.index, inplace=True)
    name_of_endtable.drop(columns="index", inplace=True)
    name_of_endtable["Points Total"] = table_with_riders["Points Total"]
    name_of_endtable.columns = table_with_riders.columns
    return name_of_endtable

    
    
"""
Filtering files including "R1" within name, returning race files
"""
results_files = os.listdir(results_path)
r = re.compile(".*R1.xml")
race_files = list(filter(r.match, results_files))
race_files_season=[]

def set_new_season():
    "Filtering files by date"
    time_format = "%Y_%m_%d_%H_%M"
    present = datetime.now()
    for race in race_files:        
        race_date = datetime.strptime(race[0:16], time_format)
        if race_date > present:
            race_files_season.append(race)
            
    book = load_workbook(statistics_folder + "team_info.xlsx")
    sheet_data = "Data"
    for sh in range(len(book.sheetnames)):
        if book.sheetnames[sh] == sheet_data:
            break
    sheet = book.active
    sheet["B7"] = season + 1
    sheet["B8"] = datetime.now()
    book.save(statistics_folder + "team_info.xlsx")
    
    


table_df = list()
table_df2 = list()
race_tracks = list()


"""
Filtering races with specific series name. 
List "Races" will contain results from all races in a season
"""
races = []
for file in race_files:
    with open(results_path + file, "r") as f:
        results = f.read()
    laps = find_name(results, "<RaceLaps>", "</RaceLaps")
    try:
        find_name(results, "lap="+laps, " point")
        try:
            find_name(results, "<Mod>" + seriesname, "</Mod>")
            races.append(file)
        except:
            pass
    except:
        pass
    
"""Dictionary assigning points (right) to position of rider in a race (left)"""    
table_points = {
    1 : 25,    2 : 20,
    3 : 16,    4 : 13,
    5 : 11,    6 : 10,
    7 : 9,    8 : 8,
    9 : 7,    10 : 6,
    11 : 5,    12 : 4,
    13 : 3,    14 : 2,
    15 : 1,    16 : 0,
    17 : 0,    18 : 0,
    19 : 0,    20 : 0,
    21 : 0,    22 : 0,
    23 : 0,    24 : 0,
    25 : 0,    26 : 0,
    27 : 0,    28 : 0,
    29 : 0,    30 : 0
    }


"""
Loop for dataframes with results. It opens every race report, text-scraping to gain drivers,
positions, assigning points and event names.
Later using function to create tables it is possible to create tables:
    -table riders with points
    -table teams with points
    -table riders with positions
"""
for race in races:
    with open(results_path + race, "r") as f:
        results = f.read()
    results_div = results.split("<Driver>")
    table_list = []
    positions_list = []
    for i in range(len(results_div)-1):
        start_table = results_div[i+1].index("<Position>")
        length_table = len("<Position>")
        #grid = results_div[i+1][(start_grid + length_grid) : (start_grid+length_grid+1)]
        position = int("".join(c for c in results_div[i+1][
            (start_table + length_table) : (start_table+length_table+2)] if c.isdigit()))

        table_list.append([find_name(results_div[i+1],"\n<Name>", "</Name>"),
                           find_name(results_div[i+1],"\n<TeamName>", "</TeamName>"),
                           find_name(results_div[i+1],"\n<CarType>", "</CarType>"),
                           table_points[int(position)]])
        positions_list.append([find_name(results_div[i+1],"\n<Name>", "</Name>"),
                           find_name(results_div[i+1],"\n<TeamName>", "</TeamName>"),
                           find_name(results_div[i+1],"\n<CarType>", "</CarType>"),
                           int(position)])
    race_tracks.append(find_name(results_div[0], "<TrackVenue>", "</TrackVenue>").split()[0])
    table_df.append(pd.DataFrame(table_list))
    table_df2.append(pd.DataFrame(positions_list))

table_full = create_table(table_df)
team_table = create_team_table(table_full)
riders_table = create_table_positions(table_full)



table_full
team_table
riders_table

"""
Opening our file with statistics and moving tables to excel tabs.
If tab does not exist, it creates it.
If tab exists, it overwrites.
"""
book = load_workbook(statistics_folder + "team_info.xlsx")
writer = pd.ExcelWriter(statistics_folder + "team_info.xlsx", engine = "openpyxl")
writer.book = book
team_table.to_excel(writer, sheet_name = "Teams - Season " + str(season))
riders_table.to_excel(writer, sheet_name = "Riders - Season" + str(season))
writer.close()

"""
Loading drivers from file with drivers (different file than results).
In this file drivers have all personal data assigned like nationality, skills, birthdate.
Firstly, we divide files basing on keyword. In this case this keyword is "Nationality", 
as every driver has assigned nationality and it is a way to separate drivers each other.
It returns index/row of word Nationality in a file.
Later we add them to common list.
"""
lookup = "Nationality"
nationality_text = []
with open(rcd_file) as h:
    for num, line in enumerate(h, 1):
        if lookup in line:
            print(num)
            nationality_text.append(num)
nationality_text=list(nationality_text)
nationality_text    

"""Adding all skills to list by moving indexes to gain whole (one) rider personal data"""
driver_skills = []
with open(rcd_file) as h:        
    for i, line in enumerate(h):
        for n in range(len(nationality_text)-1):
            if i in range(nationality_text[n]-3,nationality_text[n+1]-3):
                driver_skills.append([nationality_text[n], line.strip()])     
driver_skills

"""
Division of riders to list, inside one list there are another lists with skills.
One element of list is list with all driver data.
"""
#podzial kierowców w listę, w jednej liscie są podlisty z umiejetnosciami
driver_test=[]
j=0
n=0
driver_test.append([])
for i in range(0,len(driver_skills)):
    if driver_skills[i][0] == driver_skills[i-1][0]:
        driver_test[j].append(driver_skills[i][1])
    else:
        driver_test.append([])
        j=j+1
        driver_test[j].append(driver_skills[i][1])

driver_test



"""
grid_lap1 - difference between position at start and after first lap
table_grid_lap1 - creating list with lists (grid_lap1)

"""
table_grid_lap1 = []
for race in races:
    with open(results_path + race, "r") as f:
        results = f.read()
    results_div = results.split("<Driver>")
    grid_lap1 = list()
    #loop creating list surname-gain/lost positions after lap 1
    for i in range(len(results_div)-1):
        start=results_div[i+1].index("<Lap num=\"" + str(1) + "\" p=\"")
        length=len("<Lap num=\"" + str(1) + "\" p=\"")
        #lap1 = results_div[i+1][(start + length) : (start+length+1)]
        lap1 = int("".join(c for c in results_div[i+1][(start + length) : (start+length+2)] if c.isdigit()))


        start_grid = results_div[i+1].index("<GridPos>")
        length_grid = len("<GridPos>")
        grid = int("".join(c for c in results_div[i+1][(start_grid + length_grid) : (start_grid+length_grid+2)] if c.isdigit()))

        grid_lap1.append([find_name(results_div[i+1],"\n<Name>", "</Name>"),
                          find_name(results_div[i+1],"\n<TeamName>", "</TeamName>"),
                          find_name(results_div[i+1],"\n<CarType>", "</CarType>"),
                          int(grid)-int(lap1)])
    table_grid_lap1.append(pd.DataFrame(grid_lap1))
table_grid_lap1

"""Table with changes of positions after lap 1, summarised with all races."""
table_position_after_lap1 = create_table(table_grid_lap1)


"Loop being part of next loop (grid_df, for race in races)"
def loop_riders_grid():
    grid_stats = list()
    for i in range(len(results_div)-1):
        start_grid = results_div[i+1].index("<GridPos>")
        length_grid = len("<GridPos>")
        grid = int("".join(c for c in results_div[i+1][(start_grid + length_grid) : (start_grid+length_grid+2)] if c.isdigit()))

        grid_stats.append([find_name(results_div[i+1],"\n<Name>", "</Name>"), 
                           find_name(results_div[i+1],"\n<TeamName>", "</TeamName>"),
                           find_name(results_div[i+1],"\n<CarType>", "</CarType>"),
                           int(grid)])

    grid_df.append(pd.DataFrame(grid_stats))

"""Creating dataframe with grid, similar way like previous table, returning just start positions."""
grid_df = []
for race in races:
    with open(results_path + race, "r") as f:
        results = f.read()
    results_div = results.split("<Driver>")
    loop_riders_grid()

table_grid = create_table(grid_df)
table_grid.columns.values[-1] = "Sum Grid Pos"
table_grid


"""Returning riders from table and maximum points from table (it will be needed later)."""
list_of_riders = list(table_full.iloc[:,0])
#rider_pts = table_full[table_full.iloc[:,0]==driver_test[1][0]].iloc[0,-1]
max_pts = table_full.iloc[:,-1].max()


